#!/usr/bin/env python3

import os
import pandas as pd
import re
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment

def parse_ppad_file(file_path):
    """Parse PPAD.out file and extract TNS, Power, HPWL, Displacement."""
    data = {}
    try:
        with open(file_path, 'r') as f:
            content = f.read()

        # Extract TNS
        tns_match = re.search(r'TNS:\s*([-\d.]+)', content)
        if tns_match:
            data['TNS'] = float(tns_match.group(1))

        # Extract Power
        power_match = re.search(r'Power:\s*([\d.eE+-]+)', content)
        if power_match:
            data['Power'] = power_match.group(1)  # Keep as string

        # Extract HPWL
        hpwl_match = re.search(r'HPWL:\s*([\d.]+)', content)
        if hpwl_match:
            data['HPWL'] = float(hpwl_match.group(1))

        # Extract Displacement
        disp_match = re.search(r'Displacement:\s*([\d.]+)', content)
        if disp_match:
            data['Displacement'] = float(disp_match.group(1))

    except Exception as e:
        print(f"Error parsing {file_path}: {e}")

    return data

def parse_ppa_file(file_path):
    """Parse PPA.out file and extract WNS."""
    try:
        with open(file_path, 'r') as f:
            content = f.read()

        # Extract WNS
        wns_match = re.search(r'WNS:\s*([-\d.]+)', content)
        if wns_match:
            return float(wns_match.group(1))

    except Exception as e:
        print(f"Error parsing {file_path}: {e}")

    return None

def parse_runtime_file(file_path):
    """Parse runtime.log file and extract runtime."""
    try:
        with open(file_path, 'r') as f:
            content = f.read()

        # Extract runtime
        runtime_match = re.search(r'runtime\(s\):\s*(\d+)', content)
        if runtime_match:
            return int(runtime_match.group(1))

    except Exception as e:
        print(f"Error parsing {file_path}: {e}")

    return None

def fmt3(x):
    try:
        return round(float(x), 3)
    except (TypeError, ValueError):
        return x

def main():
    # Check for fold name argument
    if len(sys.argv) < 2:
        print("Usage: make_table_tiny.py <fold_name>")
        print("Example: make_table_tiny.py fold1")
        sys.exit(1)

    fold_name = sys.argv[1]
    output_dir = Path("./output")
    results = []

    # Get all testcase directories
    testcases = [d for d in output_dir.iterdir() if d.is_dir() and d.name != "ASAP7"]

    for testcase_dir in sorted(testcases):
        testcase_name = testcase_dir.name
        print(f"Processing {testcase_name}/{fold_name}...")

        # File paths - updated to use fold_name
        ppad_file = testcase_dir / fold_name / "PPAD.out"
        ppa_file = testcase_dir / fold_name / "PPA.out"
        runtime_file = testcase_dir / "runtime.log"

        # Initialize row data
        row_data = {"Testcase": testcase_name}

        # Parse PPAD.out data (TNS, Power, HPWL, Displacement)
        ppad_data = {}
        if ppad_file.exists():
            ppad_data = parse_ppad_file(ppad_file)
            row_data["TNS"] = (ppad_data['TNS'] if 'TNS' in ppad_data else 'N/A')
            row_data["Power"] = (ppad_data['Power'] if 'Power' in ppad_data else 'N/A')
            row_data["HPWL"] = (ppad_data['HPWL'] if 'HPWL' in ppad_data else 'N/A')
            row_data["Avg_Displacement"] = (ppad_data['Displacement'] if 'Displacement' in ppad_data else 'N/A')
        else:
            row_data["TNS"] = 'N/A'
            row_data["Power"] = 'N/A'
            row_data["HPWL"] = 'N/A'
            row_data["Avg_Displacement"] = 'N/A'

        # Parse PPA.out data (WNS)
        if ppa_file.exists():
            wns = parse_ppa_file(ppa_file)
            row_data["WNS"] = wns if wns is not None else 'N/A'
        else:
            row_data["WNS"] = 'N/A'

        # Parse runtime.log data
        if runtime_file.exists():
            runtime = parse_runtime_file(runtime_file)
            row_data["Runtime"] = runtime if runtime is not None else 'N/A'
        else:
            row_data["Runtime"] = 'N/A'

        results.append(row_data)

    # Create DataFrame with ordered columns
    column_order = [
        "Testcase",
        "WNS", "TNS", "Power", "HPWL",
        "Avg_Displacement", "Runtime"
    ]

    df = pd.DataFrame(results)
    df = df[column_order]

    # Create workbook and worksheet with simple headers
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    ws.sheet_view.showZeros = True

    # Simple single-row headers
    ws['A1'] = 'testcase'
    ws['B1'] = 'WNS'
    ws['C1'] = 'TNS'
    ws['D1'] = 'Power'
    ws['E1'] = 'HPWL'
    ws['F1'] = 'avg.disp'
    ws['G1'] = 'runtime(s)'

    # Add data starting from row 2
    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        ws[f'A{row_idx}'] = row['Testcase']
        ws[f'B{row_idx}'] = row['WNS']
        ws[f'C{row_idx}'] = row['TNS']
        ws[f'D{row_idx}'] = row['Power']
        ws[f'E{row_idx}'] = row['HPWL']

        avg_d = fmt3(row['Avg_Displacement'])
        avg_d_cell = ws[f'F{row_idx}']; avg_d_cell.value = avg_d
        if isinstance(avg_d, (int, float)): avg_d_cell.number_format = '0.000'

        ws[f'G{row_idx}'] = row['Runtime']

    # Write to Excel
    output_file = "./output/results_tiny.xlsx"
    wb.save(output_file)

    print(f"Results written to {output_file}")
    print(f"Processed {len(results)} testcases")

if __name__ == "__main__":
    main()
