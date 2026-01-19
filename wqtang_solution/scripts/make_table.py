#!/usr/bin/env python3

import os
import pandas as pd
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment

def parse_ppa_file(file_path):
    """Parse PPA file and extract WNS, TNS, Power, HPWL values."""
    data = {}
    try:
        with open(file_path, 'r') as f:
            content = f.read()

        # Extract WNS
        wns_match = re.search(r'WNS:\s*([-\d.]+)', content)
        if wns_match:
            data['WNS'] = float(wns_match.group(1))

        # Extract TNS
        tns_match = re.search(r'TNS:\s*([-\d.]+)', content)
        if tns_match:
            data['TNS'] = float(tns_match.group(1))

        # Extract Power
        power_match = re.search(r'Power:\s*([\d.eE+-]+)', content)
        if power_match:
            data['Power'] = power_match.group(1)  # Keep as string
            data['Power_float'] = float(power_match.group(1))  # For calculations

        # Extract HPWL
        hpwl_match = re.search(r'HPWL:\s*([\d.]+)', content)
        if hpwl_match:
            data['HPWL'] = float(hpwl_match.group(1))
            
    except Exception as e:
        print(f"Error parsing {file_path}: {e}")
        
    return data

def parse_best_ppad(file_path):
    """Parse best.ppad file and extract WNS, TNS, Power, HPWL, Displacement."""
    data = {}
    try:
        with open(file_path, 'r') as f:
            content = f.read()

        # Extract WNS
        wns_match = re.search(r'WNS:\s*([-\d.]+)', content)
        if wns_match:
            data['WNS'] = float(wns_match.group(1))

        # Extract TNS
        tns_match = re.search(r'TNS:\s*([-\d.]+)', content)
        if tns_match:
            data['TNS'] = float(tns_match.group(1))

        # Extract Power
        power_match = re.search(r'Power:\s*([\d.eE+-]+)', content)
        if power_match:
            data['Power'] = power_match.group(1)  # Keep as string
            data['Power_float'] = float(power_match.group(1))  # For calculations

        # Extract HPWL
        hpwl_match = re.search(r'HPWL:\s*([\d.]+)', content)
        if hpwl_match:
            data['HPWL'] = float(hpwl_match.group(1))

        # Extract Displacement
        disp_match = re.search(r'Displacement:\s*([\d.]+)', content)
        if disp_match:
            data['Displacement'] = float(disp_match.group(1))
            
    except Exception as e:
        print(f"Error parsing {file_path}: {e}")
        
    return data

def parse_best_score(file_path):
    """Parse best.score file and extract the final score."""
    try:
        with open(file_path, 'r') as f:
            content = f.read()
            
        # Extract score
        score_match = re.search(r'S\s*=\s*([\d.-]+)', content)
        if score_match:
            return float(score_match.group(1))
            
    except Exception as e:
        print(f"Error parsing {file_path}: {e}")
        
    return None

def calculate_improvement(original, optimized):
    """Calculate improvement percentage."""
    if original == 0:
        return 0.0
    return ((optimized - original) / abs(original)) * 100

def calculate_reduction(original, optimized):
    """Calculate improvement percentage."""
    if original == 0:
        return 0.0
    return ((original - optimized) / abs(original)) * 100

def fmt3(x):
    try:
        return round(float(x), 3)
    except (TypeError, ValueError):
        # 例如 'N/A' 或 None 就原樣返回
        return x
    
def main():
    output_dir = Path("./output")
    results = []
    
    # Get all testcase directories
    testcases = [d for d in output_dir.iterdir() if d.is_dir() and d.name != "ASAP7"]
    
    for testcase_dir in sorted(testcases):
        testcase_name = testcase_dir.name
        print(f"Processing {testcase_name}...")
        
        # File paths
        original_file = testcase_dir / "original.PPA.out"
        best_file = testcase_dir / "best.ppad"
        score_file = testcase_dir / "best.score"
        
        # Initialize row data
        row_data = {"Testcase": testcase_name}
        
        # Parse original data
        original_data = {}
        original_file_to_use = None
        
        if original_file.exists():
            original_file_to_use = original_file
        else:
            # Try to find it in database directory
            database_original_file = Path(f"./database/{testcase_name}/original/PPAD.out")
            if database_original_file.exists():
                original_file_to_use = database_original_file
        
        if original_file_to_use:
            original_data = parse_ppa_file(original_file_to_use)
            row_data["Original_WNS"] = (original_data['WNS'] if 'WNS' in original_data else 'N/A')
            row_data["Original_TNS"] = (original_data['TNS'] if 'TNS' in original_data else 'N/A')
            row_data["Original_Power"] = (original_data['Power'] if 'Power' in original_data else 'N/A')
            row_data["Original_HPWL"] = (original_data['HPWL'] if 'HPWL' in original_data else 'N/A')
        else:
            row_data["Original_WNS"] = 'N/A'
            row_data["Original_TNS"] = 'N/A'
            row_data["Original_Power"] = 'N/A'
            row_data["Original_HPWL"] = 'N/A'
        
        # Parse optimized data
        optimized_data = {}
        if best_file.exists():
            optimized_data = parse_best_ppad(best_file)
            row_data["Optimized_WNS"] = (optimized_data['WNS'] if 'WNS' in optimized_data else 'N/A')
            row_data["Optimized_TNS"] = (optimized_data['TNS'] if 'TNS' in optimized_data else 'N/A')
            row_data["Optimized_Power"] = (optimized_data['Power'] if 'Power' in optimized_data else 'N/A')
            row_data["Optimized_HPWL"] = (optimized_data['HPWL'] if 'HPWL' in optimized_data else 'N/A')
            row_data["Avg_Displacement"] = (optimized_data['Displacement'] if 'Displacement' in optimized_data else 'N/A')
        else:
            row_data["Optimized_WNS"] = 'N/A'
            row_data["Optimized_TNS"] = 'N/A'
            row_data["Optimized_Power"] = 'N/A'
            row_data["Optimized_HPWL"] = 'N/A'
            row_data["Avg_Displacement"] = 'N/A'
        
        # Calculate improvements
        if original_data and optimized_data:
            row_data["WNS_Improvement_%"] = calculate_improvement(original_data.get('WNS', 0), optimized_data.get('WNS', 0))
            row_data["TNS_Improvement_%"] = calculate_improvement(original_data.get('TNS', 0), optimized_data.get('TNS', 0))
            row_data["Power_Improvement_%"] = calculate_reduction(original_data.get('Power_float', 0), optimized_data.get('Power_float', 0))
            row_data["HPWL_Improvement_%"] = calculate_reduction(original_data.get('HPWL', 0), optimized_data.get('HPWL', 0))
        else:
            row_data["WNS_Improvement_%"] = 'N/A'
            row_data["TNS_Improvement_%"] = 'N/A'
            row_data["Power_Improvement_%"] = 'N/A'
            row_data["HPWL_Improvement_%"] = 'N/A'
        
        # Parse best score
        if score_file.exists():
            score = parse_best_score(score_file)
            row_data["Score"] = score if score is not None else 'N/A'
        else:
            row_data["Score"] = 'N/A'
        
        results.append(row_data)
    
    # Create DataFrame with ordered columns
    column_order = [
        "Testcase",
        "Original_WNS", "Original_TNS", "Original_Power", "Original_HPWL",
        "Optimized_WNS", "Optimized_TNS", "Optimized_Power", "Optimized_HPWL",
        "WNS_Improvement_%", "TNS_Improvement_%", "Power_Improvement_%", "HPWL_Improvement_%",
        "Avg_Displacement",
        "Score"
    ]
    
    df = pd.DataFrame(results)
    df = df[column_order]
    
    # Create workbook and worksheet manually for custom headers
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    
    ws.sheet_view.showZeros = True
    # Create merged headers
    # Row 1: Main categories
    ws['A2'] = 'testcase'
    ws['B1'] = 'original'
    ws['F1'] = 'optimized'
    ws['J1'] = 'impv (%)'
    ws['N2'] = 'avg.disp'
    ws['O2'] = 'score'

    # Row 2: Sub-categories
    #ws['A2'] = ''  # Empty for testcase
    ws['B2'] = 'WNS'
    ws['C2'] = 'TNS'
    ws['D2'] = 'Power'
    ws['E2'] = 'HPWL'
    ws['F2'] = 'WNS'
    ws['G2'] = 'TNS'
    ws['H2'] = 'Power'
    ws['I2'] = 'HPWL'
    ws['J2'] = 'WNS'
    ws['K2'] = 'TNS'
    ws['L2'] = 'Power'
    ws['M2'] = 'HPWL'
    #ws['N2'] = ''  # Empty for avg.disp
    #ws['O2'] = ''  # Empty for score

    # Merge cells for main categories
    #ws.merge_cells('A1:A2')
    #ws.merge_cells('N1:N2')
    #ws.merge_cells('O1:O2')
    ws.merge_cells('B1:E1')  # original
    ws.merge_cells('F1:I1')  # optimized
    ws.merge_cells('J1:M1')  # impv (%)

    # Set alignment for merged headers
    for cell_range in ['B1', 'F1', 'J1']:
        ws[cell_range].alignment = Alignment(horizontal='center')
    
    # Add data starting from row 3
    for row_idx, (_, row) in enumerate(df.iterrows(), start=3):
        ws[f'A{row_idx}'] = row['Testcase']
        ws[f'B{row_idx}'] = row['Original_WNS']
        ws[f'C{row_idx}'] = row['Original_TNS']
        ws[f'D{row_idx}'] = row['Original_Power']
        ws[f'E{row_idx}'] = row['Original_HPWL']
        ws[f'F{row_idx}'] = row['Optimized_WNS']
        ws[f'G{row_idx}'] = row['Optimized_TNS']
        ws[f'H{row_idx}'] = row['Optimized_Power']
        ws[f'I{row_idx}'] = row['Optimized_HPWL']

        wns_impv = fmt3(row['WNS_Improvement_%'])
        tns_impv = fmt3(row['TNS_Improvement_%'])
        power_impv = fmt3(row['Power_Improvement_%'])
        hpwl_impv = fmt3(row['HPWL_Improvement_%'])

        wns_impv_cell = ws[f'J{row_idx}']; wns_impv_cell.value = wns_impv
        tns_impv_cell = ws[f'K{row_idx}']; tns_impv_cell.value = tns_impv
        power_impv_cell = ws[f'L{row_idx}']; power_impv_cell.value = power_impv
        hpwl_impv_cell = ws[f'M{row_idx}']; hpwl_impv_cell.value = hpwl_impv

        if isinstance(wns_impv, (int, float)): wns_impv_cell.number_format = '0.000'
        if isinstance(tns_impv, (int, float)): tns_impv_cell.number_format = '0.000'
        if isinstance(power_impv, (int, float)): power_impv_cell.number_format = '0.000'
        if isinstance(hpwl_impv, (int, float)): hpwl_impv_cell.number_format = '0.000'

        avg_d = fmt3(row['Avg_Displacement'])
        avg_d_cell = ws[f'N{row_idx}']; avg_d_cell.value = avg_d
        if isinstance(avg_d, (int, float)): avg_d_cell.number_format = '0.000'

        ws[f'O{row_idx}'] = row['Score']
    
    # Write to Excel
    output_file = "./output/results.xlsx"
    wb.save(output_file)
    
    print(f"Results written to {output_file}")
    print(f"Processed {len(results)} testcases")

if __name__ == "__main__":
    main()